import csv
import json
import logging
import time
import logging.handlers

import openpyxl
import yaml
from selenium import webdriver

import config

# 定义工具类
import time

from selenium import webdriver
from appium import webdriver as app_driver
from selenium.webdriver.common.by import By



def init_logging():
    # 创建日志器
    logger = logging.getLogger()
    # 设置日志的级别
    logger.setLevel(logging.DEBUG)
    # 创建处理器
    fh = logging.handlers.TimedRotatingFileHandler(config.base_dir() + f"/log/log_{time.strftime('%Y%m%d')}.log",
                                                   when="midnight",
                                                   interval=1, backupCount=7)
    sh = logging.StreamHandler()
    fh.setLevel(logging.INFO)
    sh.setLevel(logging.INFO)
    # 创建格式器
    fmt = "%(asctime)s %(levelname)s [%(name)s] [%(filename)s(%(funcName)s:%(lineno)d)] - %(message)s"
    formatter = logging.Formatter(fmt=fmt)
    # 在处理器添加格式器
    fh.setFormatter(formatter)
    sh.setFormatter(formatter)
    # 在日志器添加处理器
    logger.addHandler(sh)
    logger.addHandler(fh)


def parse_csv(file):
    with open(file, 'r', encoding='utf-8') as f:
        data = csv.reader(f)
        test_data = []
        for i in data:
            test_data.append(i)
        del test_data[0]
        return test_data


def parse_xls(file, sheet):
    wb = openpyxl.load_workbook(file)
    ws = wb[sheet]

    rows = ws.max_row
    cols = ws.max_column

    test_data = []
    for row in range(2, rows + 1):
        test_temp = []
        for col in range(1, cols + 1):
            value = ws.cell(row, col).value
            test_temp.append(value)
        test_data.append(test_temp)
    return test_data

def parse_json(file):
    with open(file,'r',encoding='utf-8') as f:
        data = json.load(f)
        test_data = []
        for i in data:
            test_temp=[]
            for x in i.values():
                test_temp.append(x)
            test_data.append(test_temp)
        return test_data

def parse_yml(file):
    with open(file,'r',encoding='utf-8') as f:
        data = yaml.load(f)
        test_data = []
        for i in data:
            test_temp=[]
            for x in i.values():
                test_temp.append(x)
            test_data.append(test_temp)
        return test_data

if __name__ == '__main__':
    UtilsDriver().get_driver().get('https://www.baidu.com/')
    time.sleep(2)

    UtilsDriver().get_driver().get('https://www.bilibili.com/')


    # print(parse_csv(r'C:\Users\gkxox\Desktop\TPShop_UI\Data\test_data.csv'))
    # print(parse_xls(r'C:\Users\gkxox\Desktop\TPShop_UI\Data\test_data.xlsx','Sheet1'))
    # print(parse_json(r'C:\Users\gkxox\Desktop\TPShop_UI\Data\test_data.json'))
    # print(parse_yml(r'C:\Users\gkxox\Desktop\TPShop_UI\Data\test_data.yaml'))